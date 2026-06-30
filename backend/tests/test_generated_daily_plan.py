from datetime import date, time
from hashlib import sha256
from pathlib import Path

from openpyxl import load_workbook
import pytest

from app.services.daily_plan_builder import DailyPlanBuilder
from app.services.excel_exporter import export_plan_to_xlsx
from app.services.osrm_service import OSRMTable


ROOT = Path(__file__).resolve().parents[2]
WEEKLY_DIR = ROOT / "weekly planning"
SOURCE_FILE = WEEKLY_DIR / "Weekly Delivery planning W0526.xlsx"


class FakeOSRM:
    @staticmethod
    def _meters(a, b):
        # Deterministic local stand-in for OSRM in tests.
        return ((a[0] - b[0]) ** 2 + (a[1] - b[1]) ** 2) ** 0.5 * 111_000 * 1.3

    def table(self, coordinates):
        distances = []
        durations = []
        for a in coordinates:
            dist_row = []
            dur_row = []
            for b in coordinates:
                meters = self._meters(a, b)
                dist_row.append(meters)
                dur_row.append(meters / 1000 / 55 * 3600)
            distances.append(dist_row)
            durations.append(dur_row)
        return OSRMTable(durations_sec=durations, distances_m=distances)

    def route(self, coordinates):
        legs = []
        total_m = 0.0
        total_s = 0.0
        for a, b in zip(coordinates, coordinates[1:]):
            meters = self._meters(a, b)
            seconds = meters / 1000 / 55 * 3600
            total_m += meters
            total_s += seconds
            legs.append({
                "distance_m": meters,
                "distance_km": round(meters / 1000, 3),
                "duration_sec": seconds,
                "travel_min": int(round(seconds / 60)),
                "steps": [],
            })
        return {
            "distance_m": total_m,
            "total_distance_km": round(total_m / 1000, 3),
            "duration_sec": total_s,
            "total_travel_min": int(round(total_s / 60)),
            "geometry": {"type": "LineString", "coordinates": [[lon, lat] for lat, lon in coordinates]},
            "legs": legs,
        }


def _builder():
    return DailyPlanBuilder(WEEKLY_DIR, osrm=FakeOSRM())


@pytest.fixture(scope="session", autouse=True)
def setup_test_db():
    yield


def _first_assigned_stop(plan):
    for truck in plan["trucks"]:
        for trip in truck["trips"]:
            if trip["stops"]:
                return trip["stops"][0], trip
    raise AssertionError("generated plan has no assigned stops")


def _minutes(value):
    hours, minutes = str(value).split(":")[:2]
    return int(hours) * 60 + int(minutes)


def test_hos_warnings_flag_long_days():
    builder = DailyPlanBuilder(WEEKLY_DIR)
    trucks = [{
        "truck_label": "Truck X", "truck_id": 1,
        "trips": [
            {"depart_at": "03:00", "return_at": "12:00", "stops": [{"travel_min": 300}]},
            {"depart_at": "12:30", "return_at": "18:30", "stops": [{"travel_min": 260}]},
        ],
    }]
    warnings = builder._hos_warnings(trucks)
    assert len(warnings) == 1
    w = warnings[0]
    assert w["truck"] == "Truck X"
    assert w["driving_minutes"] == 560 and w["driving_overflow_minutes"] == 20
    assert w["on_duty_minutes"] == 930 and w["on_duty_overflow_minutes"] == 150


def test_hos_no_warning_for_compliant_day():
    builder = DailyPlanBuilder(WEEKLY_DIR)
    trucks = [{
        "truck_label": "T", "truck_id": 2,
        "trips": [{"depart_at": "08:00", "return_at": "12:00", "stops": [{"travel_min": 120}]}],
    }]
    assert builder._hos_warnings(trucks) == []


def test_plan_exposes_hos_warnings_list():
    plan = DailyPlanBuilder(WEEKLY_DIR).build(date(2026, 5, 26))
    assert isinstance(plan["diagnostics"]["hos_warnings"], list)


def test_plan_reports_estimated_cost_in_tnd():
    plan = DailyPlanBuilder(WEEKLY_DIR).build(date(2026, 5, 26))
    cost = plan["estimated_cost_tnd"]
    assert isinstance(cost, dict)
    assert cost["total"] > 0
    for key in ("trucks", "fuel", "driver", "underutilization", "unassigned_penalty"):
        assert key in cost
        assert cost[key] >= 0


def test_cost_config_is_injectable():
    from app.services.daily_plan_builder import CostConfig

    cheap = CostConfig(fuel_price_tnd_per_liter=0.0, driver_hourly_cost_tnd=0.0,
                       truck_dispatch_fixed_tnd=0.0, underutil_penalty_per_pos=0.0,
                       unassigned_delivery_penalty_tnd=0.0, rental_truck_per_day_tnd=0.0)
    builder = DailyPlanBuilder(WEEKLY_DIR, cost_config=cheap)
    assert builder.cost_config.fuel_price_tnd_per_liter == 0.0


def test_clock_does_not_clamp_midnight():
    result = DailyPlanBuilder._clock(1514)  # 25h 14min
    assert result != "23:59", "_clock must not clamp late returns"
    assert result == "25:14"


def test_clock_formats_normal_time():
    assert DailyPlanBuilder._clock(6 * 60 + 5) == "06:05"


def test_daily_plan_builder_assigns_real_workbook_rows():
    plan = _builder().build(date(2026, 5, 26))

    assigned_stops = [
        stop
        for truck in plan["trucks"]
        for trip in truck["trips"]
        for stop in trip["stops"]
    ]

    # Conservation is per ORIGINAL delivery, not per stop: a split delivery is
    # served across several stops but is still one considered delivery, so we
    # collapse each stop back to its parent (split_parent_id when split, else id)
    # before counting. Every considered delivery is either served or unassigned.
    def _parent(stop):
        return stop.get("split_parent_id", stop.get("id"))

    assigned_deliveries = {_parent(s) for s in assigned_stops}
    unassigned_deliveries = {_parent(s) for s in plan["unassigned"]}

    assert plan["source_file"] == SOURCE_FILE.name
    assert plan["day"] == "2026-05-26"
    assert plan["summary"]["selected_delivery_rows"] > 0
    assert len(assigned_stops) > 0
    assert (
        len(assigned_deliveries | unassigned_deliveries)
        == plan["summary"]["deliveries_considered"]
    )


def test_daily_plan_api_filters_out_unavailable_resource_payload():
    from app.routes.optimization import _sanitize_daily_plan_trucks

    trucks = _sanitize_daily_plan_trucks([
        {"truck_id": 1, "truck_label": "Ready", "capacity_positions": 14, "capacity_kg": 10000, "resource_status": "available"},
        {"truck_id": 2, "truck_label": "Broken", "capacity_positions": 14, "capacity_kg": 10000, "resource_status": "out_of_service"},
        {"truck_id": 3, "truck_label": "Paused Driver", "capacity_positions": 14, "capacity_kg": 10000, "driver_status": "En pause"},
    ])

    assert [truck["truck_id"] for truck in trucks] == [1]


def test_daily_plan_db_fleet_requires_available_truck_and_active_driver(db):
    from app.database import Base
    from app.models.camion import Camion, CamionStatus, CamionType
    from app.models.chauffeur import Chauffeur, ChauffeurStatus, PermisType
    from app.routes.optimization import _available_trucks_for_daily_plan

    Base.metadata.create_all(bind=db.get_bind(), checkfirst=True)

    active = Chauffeur(id=501, full_name="Active Driver", permis_type=PermisType.C, status=ChauffeurStatus.ACTIF)
    paused = Chauffeur(id=502, full_name="Paused Driver", permis_type=PermisType.C, status=ChauffeurStatus.CONGE)
    ready = Camion(
        id=601, plate_number="READY-601", type=CamionType.PORTEUR, capacite_kg=10000,
        max_palettes=14, status=CamionStatus.DISPONIBLE, chauffeur_defaut_id=501,
    )
    broken = Camion(
        id=602, plate_number="BROKEN-602", type=CamionType.PORTEUR, capacite_kg=10000,
        max_palettes=14, status=CamionStatus.PANNE, chauffeur_defaut_id=501,
    )
    no_driver = Camion(
        id=603, plate_number="PAUSED-603", type=CamionType.PORTEUR, capacite_kg=10000,
        max_palettes=14, status=CamionStatus.DISPONIBLE, chauffeur_defaut_id=502,
    )
    db.add_all([active, paused, ready, broken, no_driver])
    db.commit()

    fleet = _available_trucks_for_daily_plan(db)

    assert [truck["truck_id"] for truck in fleet] == [601]


def test_daily_plan_builder_does_not_parallelize_one_truck():
    plan = _builder().build(date(2026, 5, 26))

    for truck in plan["trucks"]:
        previous_trip_return = None
        for trip in truck["trips"]:
            depart = _minutes(trip["depart_at"])
            returned = _minutes(trip["return_at"])
            assert depart < returned
            if previous_trip_return is not None:
                assert depart >= previous_trip_return
            previous_trip_return = returned

            previous_stop_end = None
            for stop in trip["stops"]:
                start = _minutes(stop["etd"])
                end = _minutes(stop["eta"])
                assert start < end
                if previous_stop_end is not None:
                    assert start >= previous_stop_end
                previous_stop_end = end


def test_no_trip_exceeds_truck_capacity_or_working_hours():
    """Every trip must fit its truck by BOTH positions and gross weight, and
    must DEPART within the legal dispatch window. The binding daily constraint is
    the departure cut-off, not the return: by design a truck serving a far zone
    may drive the empty truck home in the evening (see DailyPlanConfig). Guards
    the kg-capacity enforcement and the depart-by-cutoff rule against regressions.
    """
    from app.services.daily_plan_builder import DailyPlanConfig

    cfg = DailyPlanConfig()
    earliest_depart = _minutes(cfg.early_start)   # 05:00 (long hauls may stage early)
    latest_depart = _minutes(cfg.max_depart)      # 18:00 (hard depart cut-off)
    end_of_day = 24 * 60                           # returns must still land same calendar day

    for day in (date(2026, 5, 25), date(2026, 5, 26), date(2026, 5, 28)):
        plan = _builder().build(day)
        by_id = {t["truck_id"]: t for t in plan["trucks"]}

        for truck in plan["trucks"]:
            for trip in truck["trips"]:
                load_pos = sum(float(s.get("quantity_positions") or 0) for s in trip["stops"])
                load_kg = sum(float(s.get("quantity_kg") or 0) for s in trip["stops"])
                assert load_pos <= truck["capacity_positions"], (
                    f"{day} truck {truck['truck_id']} trip {trip['trip_id']} "
                    f"overloaded: {load_pos} > {truck['capacity_positions']} positions"
                )
                assert load_kg <= truck["capacity_kg"] + 0.5, (
                    f"{day} truck {truck['truck_id']} trip {trip['trip_id']} "
                    f"overweight: {load_kg} > {truck['capacity_kg']} kg"
                )
                depart = _minutes(trip["depart_at"])
                assert earliest_depart <= depart <= latest_depart, (
                    f"{day} truck {truck['truck_id']} trip {trip['trip_id']} "
                    f"departs outside the legal window ({trip['depart_at']})"
                )
                assert _minutes(trip["return_at"]) <= end_of_day, (
                    f"{day} truck {truck['truck_id']} trip {trip['trip_id']} "
                    f"returns after midnight ({trip['return_at']})"
                )

        # Rentals are recommendations only and never enter an initial plan.
        assert all(t.get("ownership") != "RENTAL" for t in by_id.values())


def test_rental_recommendations_choose_smallest_viable_profile_and_require_approval():
    recommendations = DailyPlanBuilder._rental_recommendations(
        [
            {"id": "small", "quantity_positions": 5, "quantity_kg": 1200, "unassigned_reason": "Exceeds working hours"},
            {"id": "heavy", "quantity_positions": 10, "quantity_kg": 8000, "unassigned_reason": "No available trucks"},
            {"id": "semi", "quantity_positions": 20, "quantity_kg": 18000, "unassigned_reason": "No available trucks"},
            {"id": "unknown", "quantity_positions": 4, "quantity_kg": 500, "unassigned_reason": "Could not locate client"},
        ]
    )

    by_profile = {item["profile"]: item for item in recommendations}
    assert set(by_profile) == {"LIGHT_5_PALLET", "HEAVY_TRUCK", "SEMI_TRAILER"}
    assert by_profile["LIGHT_5_PALLET"]["delivery_ids"] == ["small"]
    assert by_profile["HEAVY_TRUCK"]["delivery_ids"] == ["heavy"]
    assert by_profile["SEMI_TRAILER"]["delivery_ids"] == ["semi"]
    assert all(item["approval_required"] is True for item in recommendations)
    assert all(item["status"] == "PROPOSED" for item in recommendations)
    assert all(item["truck"]["ownership"] == "RENTAL" for item in recommendations)


def test_export_round_trip_preserves_source_and_writes_edited_rows(tmp_path):
    source_hash = sha256(SOURCE_FILE.read_bytes()).hexdigest()
    plan = _builder().build(date(2026, 5, 26))
    edited_stop, trip = _first_assigned_stop(plan)
    edited_row = edited_stop["raw"]["excel_row_index"]

    edited_stop["etd"] = "10:15"
    edited_stop["eta"] = "10:45"
    edited_stop["status"] = "cancelled"
    trip["stops"].append({
        "id": "new-test",
        "client": "New Export Client",
        "quantity_positions": 2,
        "position_count": 2,
        "quantity_kg": 1234,
        "etd": "14:00",
        "eta": "14:30",
        "priority": "high",
        "status": "new",
        "constraints": {"required_date": plan["day"], "notes": "manual add"},
        "raw": {},
    })

    exported = export_plan_to_xlsx(SOURCE_FILE, plan, tmp_path)

    assert sha256(SOURCE_FILE.read_bytes()).hexdigest() == source_hash
    assert exported.name.startswith("Weekly Delivery planning W0526_edited_")

    workbook = load_workbook(exported, data_only=False)
    sheet = workbook["Planning"] if "Planning" in workbook.sheetnames else workbook.active

    assert sheet.cell(edited_row, 4).value == time(10, 15)
    assert sheet.cell(edited_row, 6).value == "cancelled"

    appended_row = sheet.max_row
    assert sheet.cell(appended_row, 1).value == "Tuesday"
    assert isinstance(sheet.cell(appended_row, 2).value, int)
    assert sheet.cell(appended_row, 3).value == "New Export Client"
    assert sheet.cell(appended_row, 4).value == time(14, 0)
    assert sheet.cell(appended_row, 5).value == 2
    assert sheet.cell(appended_row, 6).value == "new"
    assert sheet.cell(appended_row, 7).value == "manual add"
    assert sheet.cell(appended_row, 10).value == "high"
    assert sheet.cell(appended_row, 12).value == 1234
