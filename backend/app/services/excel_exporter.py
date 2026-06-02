"""Excel round-trip export for generated daily planning."""

from __future__ import annotations

import shutil
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from openpyxl.cell.cell import MergedCell
from openpyxl import load_workbook


FIELD_ALIASES = {
    "delivery_day": {"delivery_day", "delivery day", "day", "jour"},
    "row_number": {"n°", "n", "no", "number", "row_number", "row number"},
    "driver": {"driver", "chauffeur"},
    "vehicle": {"vehicle", "camion", "truck"},
    "etd": {"etd", "departure", "heure_depart"},
    "eta": {"eta", "arrival", "heure_arrivee"},
    "status": {"status", "statut"},
    "client": {"client", "customer"},
    "comments": {"comments", "comment", "notes"},
    "priority": {"priority", "priorite", "priorité"},
    "position_count": {"quantity", "position", "position_nbr", "position_number", "position_no", "pallets"},
    "pallet_weight_kg": {"pallet_weight"},
    "gross_weight_kg": {"gross_weight"},
    "total_gross_weight_kg": {"total_gross_weight"},
}


def export_plan_to_xlsx(source_path: Path, plan: dict[str, Any], out_dir: Path) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d-%H%M")
    out_name = f"{source_path.stem}_edited_{timestamp}{source_path.suffix}"
    out_path = out_dir / out_name
    shutil.copy2(source_path, out_path)

    workbook = load_workbook(out_path)
    sheet = workbook["Planning"] if "Planning" in workbook.sheetnames else workbook.active
    header_row, columns = _find_header(sheet)

    row_lookup = _delivery_lookup(plan)
    next_row_number = _next_row_number(sheet, columns)
    for delivery in row_lookup.values():
        if delivery.get("status") == "new":
            _append_delivery(sheet, columns, delivery, next_row_number)
            next_row_number += 1
            continue

        raw = delivery.get("raw") or {}
        excel_row = int(raw.get("excel_row_index") or 0) or None
        if excel_row is None:
            row_number = int(delivery.get("id") or 0)
            excel_row = header_row + row_number if row_number > 0 else None
        if excel_row and excel_row <= sheet.max_row:
            _write_delivery(sheet, columns, excel_row, delivery)

    workbook.save(out_path)
    return out_path


def _delivery_lookup(plan: dict[str, Any]) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    plan_day = _weekday_from_iso(plan.get("day"))
    for truck in plan.get("trucks", []):
        label = truck.get("truck_label") or f"Truck {truck.get('truck_id')}"
        for trip in truck.get("trips", []):
            for stop in trip.get("stops", []):
                raw = stop.get("raw") or {}
                result[str(stop.get("id") or id(stop))] = {
                    **stop,
                    "_vehicle": label,
                    "_driver": stop.get("constraints", {}).get("required_driver") or "",
                    "_delivery_day": raw.get("delivery_day") or stop.get("delivery_day") or plan_day,
                    "_row_number": raw.get("row_number") or stop.get("row_number"),
                }
    for stop in plan.get("unassigned", []):
        result[str(stop.get("id") or id(stop))] = stop
    return result


def _find_header(sheet) -> tuple[int, dict[str, int]]:
    best_row = 1
    best_columns: dict[str, int] = {}
    for row_idx in range(1, min(sheet.max_row, 8) + 1):
        columns: dict[str, int] = {}
        for cell in sheet[row_idx]:
            normalized = _normalize(cell.value)
            for field, aliases in FIELD_ALIASES.items():
                if normalized in aliases and field not in columns:
                    columns[field] = cell.column
        if len(columns) > len(best_columns):
            best_row = row_idx
            best_columns = columns
    return best_row, best_columns


def _write_delivery(sheet, columns: dict[str, int], row_idx: int, delivery: dict[str, Any]) -> None:
    raw = delivery.get("raw") or {}
    constraints = delivery.get("constraints") or {}
    values = {
        "delivery_day": delivery.get("_delivery_day") or raw.get("delivery_day") or delivery.get("delivery_day"),
        "row_number": delivery.get("_row_number") or raw.get("row_number") or delivery.get("row_number"),
        "driver": delivery.get("_driver") or delivery.get("constraints", {}).get("required_driver") or "",
        "vehicle": delivery.get("_vehicle") or "",
        "etd": delivery.get("etd"),
        "eta": delivery.get("eta"),
        "status": delivery.get("status"),
        "client": delivery.get("client"),
        "comments": constraints.get("notes") or constraints.get("comment_constraint") or raw.get("notes"),
        "priority": delivery.get("priority") or raw.get("priority"),
        "position_count": delivery.get("quantity_positions") or delivery.get("position_count"),
        "pallet_weight_kg": delivery.get("pallet_weight_kg") or raw.get("pallet_weight_kg"),
        "gross_weight_kg": delivery.get("gross_weight_kg") or raw.get("gross_weight_kg") or delivery.get("quantity_kg"),
        "total_gross_weight_kg": delivery.get("total_gross_weight_kg") or raw.get("total_gross_weight_kg"),
    }
    for field, value in values.items():
        col = columns.get(field)
        if col and value is not None:
            cell = sheet.cell(row=row_idx, column=col)
            if isinstance(cell, MergedCell):
                continue
            cell.value = _coerce_cell_value(field, value)


def _append_delivery(sheet, columns: dict[str, int], delivery: dict[str, Any], row_number: int) -> None:
    row_idx = sheet.max_row + 1
    _copy_row_style(sheet, row_idx - 1, row_idx)
    if not delivery.get("_row_number"):
        delivery = {**delivery, "_row_number": row_number}
    _write_delivery(sheet, columns, row_idx, delivery)


def _next_row_number(sheet, columns: dict[str, int]) -> int:
    col = columns.get("row_number")
    if not col:
        return sheet.max_row
    numbers = []
    for row_idx in range(1, sheet.max_row + 1):
        value = sheet.cell(row=row_idx, column=col).value
        try:
            numbers.append(int(value))
        except (TypeError, ValueError):
            continue
    return (max(numbers) + 1) if numbers else 1


def _copy_row_style(sheet, source_row: int, target_row: int) -> None:
    if source_row <= 0:
        return
    for col_idx in range(1, sheet.max_column + 1):
        source = sheet.cell(row=source_row, column=col_idx)
        target = sheet.cell(row=target_row, column=col_idx)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)


def _coerce_cell_value(field: str, value: Any) -> Any:
    if field not in {"etd", "eta"}:
        return value
    if hasattr(value, "hour") and hasattr(value, "minute"):
        return value
    text = str(value).strip()
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(text, fmt).time()
        except ValueError:
            pass
    return value


def _weekday_from_iso(value: Optional[Any]) -> Optional[str]:
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value)).strftime("%A")
    except ValueError:
        return None


def _normalize(value: Optional[Any]) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower().replace("\xa0", " ")
    text = text.replace("-", " ").replace(".", " ")
    return "_".join(text.split())
